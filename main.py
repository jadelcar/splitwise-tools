import uvicorn

from fastapi import FastAPI, HTTPException, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.requests import Request
from fastapi.testclient import TestClient

import logging
from fastapi.logger import logger

from sqlalchemy.orm import Session

import starlette.status as status
from starlette.middleware import Middleware
from starlette.middleware.sessions import SessionMiddleware
# from starlette.middleware.proxy_headers import ProxyHeadersMiddleware


from splitwise import Splitwise
from splitwise.expense import Expense
from splitwise.user import ExpenseUser

import os
from decimal import *
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

# Internal imports

from core.constants import *
from core.config.settings import get_settings
from api.routes import auth, expense_upload
from core.exceptions import handlers
from core.templates import templates
from services.helpers.expense_utils import *

# from db.database import engine, SessionLocal, database.get_db
from db import crud, models, schemas, database

settings = get_settings()

# Set up FastAPI
app = FastAPI(root_path=settings.ROOT_PATH, title="Splitwise Tools", version="1.0.0")

app.add_middleware(
    SessionMiddleware,
    secret_key=settings.SECRET_KEY,
    same_site="lax",      # IMPORTANT
    https_only=False,     # True only if HTTPS
)

# app.add_middleware(ProxyHeadersMiddleware, trusted_hosts="*")

# Include routers
app.include_router(auth.router)
app.include_router(expense_upload.router)

# Configure templating
app.mount("/static", StaticFiles(directory="static"), name="static")

CONSUMER_KEY = settings.CONSUMER_KEY
CONSUMER_SECRET = settings.CONSUMER_SECRET

logging.basicConfig(level=logging.DEBUG)
logger.setLevel(logging.DEBUG)

# Configure database
models.Base.metadata.create_all(bind = database.engine) # Creates DB if not yet created


"""       ----------           PATHS       -----------            """

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    """Return home page """
    return templates.TemplateResponse("home.html", {"request": request})


"""       ----------           Create data       -----------            """
@app.get("/create_upload", name = "create_upload", response_class=HTMLResponse)
def create_upload(request: Request, db: Session = Depends(database.get_db)):
    """Create a new upload"""
    sObj = auth.get_access_token(request)
    current_user = sObj.getCurrentUser().getId()
    new_upload = crud.create_upload(db, creator_user_id = current_user)
    return templates.TemplateResponse("home.html", {"request": request, "new_upload" : new_upload})

@app.post("/create_expense", name = "create_expense", response_class=HTMLResponse)
def create_expense(request: Request, db: Session = Depends(database.get_db)):
    """Create a new expense"""
    sObj = auth.get_access_token(request)
    current_user = sObj.getCurrentUser().getId()
    crud.create_expense(db, creator_user_id = current_user)
    return templates.TemplateResponse("home.html", {"request": request})


"""       ----------           Retrieve data       -----------            """

@app.get("/uploads", name = "uploads", response_class=HTMLResponse)
def get_uploads_by_id(request: Request, db: Session = Depends(database.get_db)):
    """Get uploads of the current user logged in the app, using it's user ID"""
    sObj = auth.get_access_token(request)
    current_user_id = sObj.getCurrentUser().getId()
    uploads = crud.get_uploads_by_sw_user_id(db, sw_user_id=current_user_id)
    return templates.TemplateResponse("uploads.html", {"request": request, "uploads" : uploads})

@app.get("/groups", name = "groups", response_class=HTMLResponse)
def get_groups_by_id(request: Request):
    """Get groups of the current user logged in the app, using it's app user ID"""
    sObj = auth.get_access_token(request)
    groups = sObj.getGroups()
    return templates.TemplateResponse("groups.html", {"request": request, "groups" : groups})

@app.get("/group_template/{group_id}", name = "download_group_template")
def get_template_by_group_id(request: Request, group_id: int):
    """Create a template excel for a group"""
    sObj = auth.get_access_token(request)
    group = sObj.getGroup(group_id)
    members = group.members
    
    # Define headers    
    expense_headers = ["ID","Description","Date","Amount","Currency","Paid by","All equal","Split type"]
    members_headers = ["Name", "ID"]
    
    # List of member names
    member_names = []
    for member in members:
        member_last_name = "" if member.last_name == None else member.last_name
        member_full_name = f"{member.first_name} {member_last_name}".strip()
        member_names.append(member_full_name)

    # For repeated names, enumerate them 1-N (e.g. Javier 1, Javier 2...)
    member_names_enum = []
    for i, name in enumerate(member_names): 
        count = member_names.count(name)
        if count > 1:
            member_names_enum.append(name + " " + str(member_names[0:i + 1].count(name)))
        else:
            member_names_enum.append(name)

    # Append to headers
    expense_headers += [f"_{name}" for name in member_names_enum] # add prefix "_"
    
    # Create and configure excel file
    wb = Workbook()
    expenses_ws = wb.active
    expenses_ws.title = "Expenses"
    expenses_ws.sheet_properties.tabColor = "00cc00"

    members_ws = wb.create_sheet(title="Members")
    members_ws.sheet_properties.tabColor = "009933"

    # Create headers
    for sheet_headers, sheet in zip([expense_headers, members_headers], [expenses_ws, members_ws]):
        for c, header in enumerate(sheet_headers, start=1):
            cell = sheet.cell(row=1, column=c)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thick', color="000000"))

    # Add member data in 'members' sheet
    for r, (member, member_name) in enumerate(zip(members, member_names_enum), start = 2):
        members_ws.cell(row = r, column = 1, value = member_name)
        members_ws.cell(row = r, column = 2, value = member.id)

    # Save and return file
    # Create the directory for group templates if it doesn't exist
    group_templates_dir = "static/assets/group_templates"
    if not os.path.exists(group_templates_dir):
        os.makedirs(group_templates_dir)
    file_path = f"{group_templates_dir}/{group_id}.xlsx"
    wb.save(file_path)

    headers = {'Content-Disposition': f'attachment; filename="template-{group.name}.xlsx"'}
    return FileResponse(file_path, headers=headers, media_type = "application/vnd.ms-excel")


"""       ----------           User registration       -----------            """

@app.get("/register", name = "register", response_class = HTMLResponse)
def register_show_form(request: Request):
    """ Show user a page for registering"""
    return templates.TemplateResponse("register.html", {"request": request})

@app.post("/register_submit", name = "register_submit")
def register(request: Request, username: str = Form(), password: str = Form(), confirmation: str = Form(), db: Session = Depends(database.get_db)):
    """ Register user using data in their registration form"""
    
    # Obtain user's data entered in the form
    username = request.get("username")
    password = request.get("password")
    confirmation = request.get("confirmation") 
    
    # Check if the username is being used
    db_user_byusername = crud.get_user_by_username(db, username=username)
    if db_user_byusername:
        return handlers.apology("That username is taken...", request, 400)

    # Validate data
    elif username=="" or password=="" or confirmation=="":
        # If at least one of the above are missing
        return handlers.apology("You didn't complete all the fields, right?", request, 400)
    elif password != confirmation :
        return handlers.apology("Passwords don't match bro", request, 400)

    # Create user in database
    try:
        new_user = schemas.UserCreate(username = username, password = password)
        crud.create_user(db, new_user)
        new_db_user = crud.get_user_by_username(db, username)
        request.session['user_id'] = new_db_user.id # Store the user id in session 
        request.session['username'] = new_db_user.username # Store the user id in session 
    except:
        return handlers.apology("Could not insert you in the database", request, 400)

    # Redirect to home
    return RedirectResponse('/', status_code=status.HTTP_302_FOUND) # See https://stackoverflow.com/a/65512571/19667698

@app.get("user/{username}", name = "get_user_byusername")
def get_user_byusername(username: int, db: Session = Depends(database.get_db)):
    """Get user by its username"""
    user = crud.get_user_by_username(db, username = username).first()
    return user

""" Create users and groups in database"""
@app.post("/users/", name = "create_user", response_model=schemas.User)
def create_user(user: schemas.UserCreate, db: Session = Depends(database.get_db)):
    """Create a new user in database"""
    db_user = crud.get_user_by_email(db, email=user.email)
    if db_user: #If db_user exists (i.e. the search by email return something)
        raise HTTPException(status_code=400, detail="Email already registered")
    return crud.create_user(db=db, user=user)





# Add test client and first test
client = TestClient(app)


def test_read_main():
    response = client.get("/")
    assert response.status_code == 200
    assert response.json() == {"msg": "Hello World"}


if __name__ == "__main__":
    settings = get_settings()
    uvicorn.run(
        "main:app",
        host=settings.APP_HOST,
        port=settings.APP_PORT,
        reload=True,
        reload_dirs=["db",""]
    )
